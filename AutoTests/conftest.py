import pytest
import configparser
import os
from openpyxl import load_workbook

import selenium
from selenium import webdriver

@pytest.fixture(scope="session")
def config():
    cfg = configparser.ConfigParser()
    path = os.path.join(os.path.dirname(__file__), "config.ini")
    cfg.read(path)
    return cfg


@pytest.fixture
def driver():
  #Init selenium webdriver
    driver = webdriver.Chrome()
    driver.maximize_window()
    yield driver
    driver.quit()


def read_excel_sheet(file_path: str, sheet_name: str):
    try:
        workbook = load_workbook(file_path)
    except Exception:
        return []
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        workbook.close()
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for r in rows[1:]:
        row_map = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        data.append(row_map)
    workbook.close()
    return data


@pytest.fixture()
def excel_reader():
    file_path = os.path.join(os.path.dirname("./TestData"), "testdata.xlsx")

    def _reader(file_path: str, sheet_name: str):
        return read_excel_sheet(file_path, sheet_name)

    return _reader
